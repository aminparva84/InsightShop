#!/usr/bin/env python3
"""
Generate InsightShop_User_Capabilities.xlsx from the canonical list of user actions.
Run from project root: python scripts/generate_user_capabilities_excel.py
"""
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Install openpyxl: pip install openpyxl")
    sys.exit(1)

# Canonical list: Area, Action, Route/Location, User type (Guest / Logged-in / Superadmin), Notes
ROWS = [
    ("Area", "Action", "Route / Location", "User Type", "Notes"),
    ("Browsing & Discovery", "View home hero and seasonal banners (Spring, Summer, Fall, Winter)", "/", "All", ""),
    ("Browsing & Discovery", "Browse by category (Men, Women, Kids, Pants, Shirts, etc.)", "/", "All", "Links to /products with filters"),
    ("Browsing & Discovery", "View featured products on home", "/", "All", ""),
    ("Browsing & Discovery", "Open AI chat from hero", "/", "All", ""),
    ("Browsing & Discovery", "Browse all products with pagination and Load More", "/products", "All", ""),
    ("Browsing & Discovery", "Filter by category, color, size, fabric, season, clothing category, price, search, on sale", "/products", "All", ""),
    ("Browsing & Discovery", "View AI Dashboard tab (AI-recommended products)", "/products", "All", "When navigated from AI chat"),
    ("Browsing & Discovery", "Select products to compare and go to Compare page", "/products", "All", ""),
    ("Browsing & Discovery", "View product detail (info, images, price, description, reviews)", "/products/:id", "All", ""),
    ("Browsing & Discovery", "Select color and size on product page", "/products/:id", "All", "When available"),
    ("Browsing & Discovery", "See stock and low-stock notices", "/products/:id", "All", ""),
    ("Shopping", "Add to cart (from product detail or cards) with color/size/quantity", "Product detail, Product cards", "All", "Guest or logged-in"),
    ("Shopping", "View cart", "/cart", "All", ""),
    ("Shopping", "Update quantity, color, or size in cart", "/cart", "All", ""),
    ("Shopping", "Remove item from cart or clear cart", "/cart", "All", ""),
    ("Shopping", "View You might also like / related product suggestions", "/cart", "All", ""),
    ("Shopping", "View matching pairs suggestions (outfit suggestions)", "/cart", "All", ""),
    ("Shopping", "Proceed to checkout", "/cart", "All", "Redirects to login if not logged in"),
    ("Shopping", "Enter shipping address (name, address, city, state, zip, country, phone)", "/checkout", "Logged-in", "Required for checkout"),
    ("Shopping", "Choose shipping method (rates from API)", "/checkout", "Logged-in", ""),
    ("Shopping", "Choose payment method (Stripe, JPMorgan, or Chase)", "/checkout", "Logged-in", ""),
    ("Shopping", "Option to create account at checkout (email + password)", "/checkout", "Guest", ""),
    ("Shopping", "Place order and complete payment", "/checkout", "Logged-in", ""),
    ("Shopping", "View order confirmation after checkout", "/order-confirmation", "Logged-in", "By orderId in URL"),
    ("Account & Auth", "Register with email, password, first name, last name", "/register", "Guest", ""),
    ("Account & Auth", "Log in with email/password", "/login", "Guest", ""),
    ("Account & Auth", "Log in with Google Sign-In", "/login", "Guest", ""),
    ("Account & Auth", "Redirect after login (e.g. to checkout or previous page)", "/login", "Guest", ""),
    ("Account & Auth", "Verify email via activation link", "/activation", "Guest", "?verify=..."),
    ("Account & Auth", "Log out", "Navbar", "Logged-in", ""),
    ("Member Area", "View dashboard stats (total orders, pending, total spent, completed)", "/members", "Logged-in", "Dashboard tab"),
    ("Member Area", "View recent orders list", "/members", "Logged-in", ""),
    ("Member Area", "View You might also like suggestions", "/members", "Logged-in", ""),
    ("Member Area", "View wishlist (inline tab)", "/members", "Logged-in", "Link to full /wishlist"),
    ("Member Area", "View all orders with items, totals, tax, shipping, address", "/members", "Logged-in", "Orders tab"),
    ("Member Area", "View payments info", "/members", "Logged-in", "Payments tab"),
    ("Wishlist", "View saved products", "/wishlist", "Logged-in", ""),
    ("Wishlist", "See Price drop and Back in stock badges", "/wishlist", "Logged-in", ""),
    ("Wishlist", "Add or remove from wishlist (heart on product cards or detail)", "Product cards, Product detail", "Logged-in", ""),
    ("Compare", "Compare multiple products side by side", "/compare?ids=...", "All", "From Products page selection"),
    ("Compare", "Add compared products to cart with color/size", "/compare", "All", ""),
    ("AI Assistant", "Open floating AI assistant (contextual hints)", "Site-wide (except Admin)", "All", ""),
    ("AI Assistant", "Ask for product recommendations in natural language", "AI Chat", "All", "e.g. blue jacket under $100"),
    ("AI Assistant", "Get product suggestions with IDs; ask to compare them", "AI Chat", "All", ""),
    ("AI Assistant", "Navigate to product pages or Compare from chat", "AI Chat", "All", ""),
    ("AI Assistant", "On Home: AI can update featured products section", "AI Chat on /", "All", ""),
    ("AI Assistant", "On Products: AI results in AI Dashboard tab", "AI Chat on /products", "All", ""),
    ("Reviews & Engagement", "Read product reviews", "Product detail", "All", ""),
    ("Reviews & Engagement", "Submit a review (rating 1–5 + comment)", "Product detail", "All", "Name required if not logged in"),
    ("Info & Support", "Read About (Our Story, mission, values)", "/about", "All", ""),
    ("Info & Support", "Send a message (name, email, optional order number, message)", "/contact", "All", ""),
    ("Info & Support", "Read shipping information", "/shipping", "All", ""),
    ("Info & Support", "Read returns and refund policy", "/returns", "All", ""),
    ("Admin", "View admin dashboard and quick links", "/admin", "Superadmin", ""),
    ("Admin", "View and edit Fashion KB (colors, style, occasions, fabrics, etc.)", "/admin", "Superadmin", "Fashion tab"),
    ("Admin", "View payment logs", "/admin", "Superadmin", "Payment logs tab"),
    ("Admin", "List users; view user details; reset user password", "/admin", "Superadmin", "Users tab"),
    ("Admin", "Create and edit sales (name, type, discount, dates, filters)", "/admin", "Superadmin", "Sales tab"),
    ("Admin", "List, add, and edit products (details, stock, sale, images)", "/admin", "Superadmin", "Products tab"),
    ("Admin", "Sync products to vector DB for AI search", "/admin", "Superadmin", "Products tab"),
    ("Admin", "View and manage product reviews", "/admin", "Superadmin", "Reviews tab"),
    ("Admin", "View orders and order detail", "/admin", "Superadmin", "Orders tab"),
    ("Admin", "View active carts (guest and user)", "/admin", "Superadmin", "Carts tab"),
    ("Admin", "View site statistics", "/admin", "Superadmin", "Statistics tab"),
    ("Admin", "Configure AI providers (API keys, test, default provider)", "/admin", "Superadmin", "AI Assistant tab"),
]

OUTPUT_DIR = "docs"
OUTPUT_FILE = "InsightShop_User_Capabilities.xlsx"


def main():
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    out_path = os.path.join(root, OUTPUT_DIR, OUTPUT_FILE)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "User Capabilities"

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row_idx, row in enumerate(ROWS, 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 36
    ws.freeze_panes = "A2"

    wb.save(out_path)
    print(f"Created: {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
